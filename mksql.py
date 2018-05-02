# python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import numbers
from datetime import datetime
import re
import sys


def msql(wb_name):
    wb = load_workbook('./uploadfile/' + wb_name)
    ws = wb.active
    resultlist = []
    
    tgvalues = ['09:00-20:59' , '*' , '01' , 'jan']
    tgvalues[3] = datetime.now().strftime("%b").lower()
    for row in ws.rows:
        for cell in row:
            flag = 0
            cellv = str(cell.value)
    #        print(type(cellv))
            index = cellv.find("月")
    #        t2str = cell.value.strftime('%Y-%m-%d %H:%M:%S')
            if cell.value == None:
                flag = 1
                continue
            elif index != -1:
                continue
            elif cell.column == "B":
    #            cell.number_format = numbers.FORMAT_DATE_DDMMYY
    #            print(cell.value)
    #            print(type(cell.value))
                t2str = cell.value.strftime('%Y-%m-%d %H:%M:%S')
    #            print(t2str)
    #            print(type(t2str))
                if t2str == "1899-12-31 00:00:00":
                    t2str = "1900-01-01 00:00:00"
                #print(cell.value)
                if t2str == "1900-01-01 00:00:00":
                    tstr = '01'
                else:
                    tstr = cell.value.strftime('%d')
                tgvalues[2]  = tstr
            elif cell.column == "D":
                dst_tmp = re.sub(r'21:00','20:59',cell.value)
                dst = dst_tmp.replace('～','-')
    #            print(dst)
                tgvalues[0] = dst
    #    print(type(tgvalues[2]))
    #    print(tgvalues)
        if flag == 1:
            continue
        if tgvalues[0] == "休み":
            continue
        tgstr = '|'.join(tgvalues)
        sqlstr1 = "INSERT INTO asterisk.timegroups_details (timegroupid,time) VALUES (6,'"
        sqlstr2 = "');"
        sqlstr_comb = sqlstr1 + tgstr + sqlstr2
    #    print(type(sqlstr_comb))
        #print(sqlstr_comb)
        resultlist.append(sqlstr_comb)
    return resultlist
