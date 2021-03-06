# python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import numbers
from datetime import datetime
from dateutil.relativedelta import relativedelta
import re
import sys
import MySQLdb
import json




def msql(wb_name):
    wb = load_workbook('./uploadfile/' + wb_name)
    ws = wb.active
    resultlist = []
    
    tgvalues = ['09:00-20:59' , '*' , '1' , 'jan']
    #tgvalues[3] = datetime.now().strftime("%b").lower()
    today = datetime.today()
    next_month = today + relativedelta(months=1)
    tgvalues[3] = next_month.strftime('%b').lower()
    for row in ws.rows:
        for cell in row:
            flag = 0
            cellv = str(cell.value)
    #        print(type(cellv))
            index = cellv.find("月")
    #        t2str = cell.value.strftime('%Y-%m-%d %H:%M:%S')
            if cell.value == None:
                flag = 1
                continue
            elif index != -1:
                continue
            elif cell.column == "B":
    #            cell.number_format = numbers.FORMAT_DATE_DDMMYY
    #            print(cell.value)
    #            print(type(cell.value))
                t2str = cell.value.strftime('%Y-%m-%d %H:%M:%S')
    #            print(t2str)
    #            print(type(t2str))
                if t2str == "1899-12-31 00:00:00":
                    t2str = "1900-1-1 00:00:00"
                #print(cell.value)
                if t2str == "1900-1-1 00:00:00":
                    tstr = '1'
                else:
                    tstr = cell.value.strftime('%-d')
                tgvalues[2]  = tstr
            elif cell.column == "D":
                dst_tmp = re.sub(r'21:00','20:59',cell.value)
                dst = dst_tmp.replace('～','-')
    #            print(dst)
                tgvalues[0] = dst
    #    print(type(tgvalues[2]))
    #    print(tgvalues)
        if flag == 1:
            continue
        if tgvalues[0] == "休み":
            continue
        tgstr = '|'.join(tgvalues)

        # 202.78.218.161(キミネット様ホスト)
        sqlstr1 = "INSERT INTO asterisk.timegroups_details (timegroupid,time) VALUES (1,'"
        # wm0104.fonex.jp(検証用)
        #sqlstr1 = "INSERT INTO asterisk.timegroups_details (timegroupid,time) VALUES (6,'"

        sqlstr2 = "');"
        sqlstr_comb = sqlstr1 + tgstr + sqlstr2
        #print(sqlstr_comb)
        resultlist.append(sqlstr_comb)

    query = resultlist
    resultlist2 = []

    ## DB接続準備
    # 202.78.218.161(キミネット様ホスト)
    conn = MySQLdb.connect(host="202.78.218.161", port=3306, user="ruby", password="passw0rd", database="asterisk")
    # wm0104.fonex.jp(検証用)
    #conn = MySQLdb.connect(host="wm0104.fonex.jp", port=3306, user="ruby", password="passw0rd", database="asterisk")

    cur = conn.cursor()

    for query_set in query:
      print(query_set)
      cur.execute(query_set)
      resultlist2.append(cur.fetchall())

    ## DB接続解除
    cur.close()
    conn.close()
    json.dumps(resultlist2)

    return resultlist
